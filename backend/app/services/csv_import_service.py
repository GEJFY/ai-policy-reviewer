"""CSV/Excel import service for terms, check items, and writing rules."""

import csv
import io

from openpyxl import load_workbook


def read_import_file(
    file_bytes: bytes, filename: str
) -> tuple[list[dict[str, str]], list[str]]:
    """Read CSV or Excel file and return rows as list of dicts.

    Returns:
        Tuple of (rows, errors)
    """
    errors: list[str] = []
    rows: list[dict[str, str]] = []

    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""

    if ext == "csv":
        rows, errors = _read_csv(file_bytes)
    elif ext in ("xlsx", "xls"):
        rows, errors = _read_excel(file_bytes)
    else:
        errors.append(f"Unsupported file type: .{ext}")

    return rows, errors


def _read_csv(file_bytes: bytes) -> tuple[list[dict[str, str]], list[str]]:
    """Read CSV file bytes into list of dicts."""
    errors: list[str] = []
    rows: list[dict[str, str]] = []

    # Try UTF-8 first, then shift-jis
    text = None
    for encoding in ["utf-8-sig", "utf-8", "shift_jis", "cp932"]:
        try:
            text = file_bytes.decode(encoding)
            break
        except (UnicodeDecodeError, LookupError):
            continue

    if text is None:
        errors.append("ファイルのエンコーディングを判定できませんでした")
        return rows, errors

    reader = csv.DictReader(io.StringIO(text))
    for i, row in enumerate(reader, start=2):
        rows.append({k.strip(): (v or "").strip() for k, v in row.items() if k})

    return rows, errors


def _read_excel(file_bytes: bytes) -> tuple[list[dict[str, str]], list[str]]:
    """Read Excel file bytes into list of dicts."""
    errors: list[str] = []
    rows: list[dict[str, str]] = []

    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True)
        ws = wb.active
        if ws is None:
            errors.append("Excelファイルにシートがありません")
            return rows, errors

        row_iter = ws.iter_rows(values_only=True)
        header_row = next(row_iter, None)
        if header_row is None:
            errors.append("ヘッダー行がありません")
            return rows, errors

        headers = [str(h).strip() if h else "" for h in header_row]

        for row_data in row_iter:
            row_dict: dict[str, str] = {}
            for j, val in enumerate(row_data):
                if j < len(headers) and headers[j]:
                    row_dict[headers[j]] = str(val).strip() if val is not None else ""
            if any(v for v in row_dict.values()):
                rows.append(row_dict)

        wb.close()
    except Exception as e:
        errors.append(f"Excelファイルの読み込みに失敗しました: {e}")

    return rows, errors


def validate_required_columns(
    rows: list[dict[str, str]],
    required: list[str],
) -> list[str]:
    """Check that all required columns exist."""
    if not rows:
        return ["データ行がありません"]

    available = set(rows[0].keys())
    missing = [col for col in required if col not in available]
    if missing:
        return [f"必須列が不足しています: {', '.join(missing)}"]
    return []


def generate_csv_template(headers: list[str], sample_row: list[str]) -> bytes:
    """Generate a CSV template with headers and sample data."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    writer.writerow(sample_row)
    # BOM for Excel compatibility
    return ("\ufeff" + output.getvalue()).encode("utf-8")


# --- Template definitions ---

TERM_HEADERS = ["term", "aliases", "definition", "category", "usage_note"]
TERM_SAMPLE = [
    "従業員",
    '["社員", "スタッフ"]',
    "会社と雇用契約を締結した者",
    "人事",
    "正社員・契約社員を含む",
]

CHECK_ITEM_HEADERS = [
    "name",
    "category",
    "description",
    "severity",
    "prompt_template",
    "is_active",
]
CHECK_ITEM_SAMPLE = [
    "用語統一チェック",
    "TERMINOLOGY",
    "文書内の用語が統一されているか確認します",
    "MEDIUM",
    "",
    "true",
]

WRITING_RULE_HEADERS = [
    "name",
    "rule_type",
    "pattern",
    "correct_form",
    "example_bad",
    "example_good",
    "is_active",
]
WRITING_RULE_SAMPLE = [
    "敬体統一",
    "STYLE",
    "「である」調と「です・ます」調の混在",
    "「です・ます」調に統一",
    "業務を遂行する。",
    "業務を遂行します。",
    "true",
]
