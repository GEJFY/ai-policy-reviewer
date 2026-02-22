"""Excel file parser for extracting text from .xlsx/.xls files."""

from app.core.logging_config import get_logger

logger = get_logger(__name__)


def extract_text_from_excel(file_path: str) -> str:
    """Extract all text content from an Excel file.

    Reads all sheets and concatenates cell values into a single text.
    """
    import openpyxl

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    except Exception as e:
        logger.error(f"Failed to open Excel file: {file_path}, error={e}")
        raise

    parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        parts.append(f"【シート: {sheet_name}】")
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            line = "\t".join(cells).strip()
            if line:
                parts.append(line)
        parts.append("")

    wb.close()
    return "\n".join(parts).strip()
