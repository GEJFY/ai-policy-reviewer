"""API endpoint for exporting review results to Excel."""

import io
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy.orm import Session

from app.db.database import get_db
from app.models.document import Document
from app.models.review import Review, ReviewFinding

router = APIRouter(prefix="/api/v1/reviews", tags=["Export"])

# スタイル定数
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

_SEVERITY_FILLS = {
    "HIGH": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "MEDIUM": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "LOW": PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),
}

_STATUS_LABELS = {
    "PENDING": "未対応",
    "APPROVED": "承認",
    "REJECTED": "却下",
    "DEFERRED": "保留",
}


@router.get("/{review_id}/export")
async def export_review_excel(review_id: int, db: Session = Depends(get_db)):
    """
    レビュー結果をExcelファイルとしてエクスポートする。

    Sheet1: レビュー概要（文書情報・統計サマリー）
    Sheet2: 指摘事項一覧（全findingsの詳細）
    """
    review = db.query(Review).filter(Review.id == review_id).first()
    if not review:
        raise HTTPException(status_code=404, detail="Review not found")

    document = db.query(Document).filter(Document.id == review.document_id).first()

    findings = (
        db.query(ReviewFinding)
        .filter(ReviewFinding.review_id == review_id)
        .order_by(ReviewFinding.severity.desc(), ReviewFinding.created_at)
        .all()
    )

    wb = _build_workbook(review, document, findings)

    # ストリーミングレスポンスとして返却
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    doc_title = document.title if document else f"review_{review_id}"
    # ファイル名に使えない文字を除去
    safe_title = doc_title.replace("/", "_").replace("\\", "_").replace(":", "_")
    filename = f"{safe_title}_レビュー結果.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{_encode_filename(filename)}"
        },
    )


def _encode_filename(filename: str) -> str:
    """RFC 5987形式でファイル名をエンコードする。"""
    from urllib.parse import quote

    return quote(filename, safe="")


def _build_workbook(
    review: Review,
    document: Document | None,
    findings: list[ReviewFinding],
) -> Workbook:
    """レビュー結果のExcelワークブックを構築する。"""
    wb = Workbook()

    _build_summary_sheet(wb, review, document, findings)
    _build_findings_sheet(wb, findings)

    return wb


def _build_summary_sheet(
    wb: Workbook,
    review: Review,
    document: Document | None,
    findings: list[ReviewFinding],
) -> None:
    """Sheet1: レビュー概要シートを構築する。"""
    ws = wb.active
    ws.title = "レビュー概要"

    # 列幅設定
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 40

    # タイトル
    ws.merge_cells("A1:B1")
    title_cell = ws["A1"]
    title_cell.value = "レビュー結果レポート"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    # 基本情報
    rows = [
        ("文書名", document.title if document else "不明"),
        ("レビューID", review.id),
        ("ステータス", review.status),
        ("作成日時", _format_datetime(review.created_at)),
        (
            "完了日時",
            _format_datetime(review.completed_at) if review.completed_at else "未完了",
        ),
    ]

    row_num = 3
    for label, value in rows:
        label_cell = ws.cell(row=row_num, column=1, value=label)
        label_cell.font = Font(bold=True)
        label_cell.border = _THIN_BORDER
        value_cell = ws.cell(row=row_num, column=2, value=value)
        value_cell.border = _THIN_BORDER
        row_num += 1

    # 統計サマリー
    row_num += 1
    ws.merge_cells(f"A{row_num}:B{row_num}")
    section_cell = ws.cell(row=row_num, column=1, value="指摘事項サマリー")
    section_cell.font = Font(bold=True, size=12)
    row_num += 1

    high_count = sum(1 for f in findings if f.severity == "HIGH")
    medium_count = sum(1 for f in findings if f.severity == "MEDIUM")
    low_count = sum(1 for f in findings if f.severity == "LOW")
    pending_count = sum(1 for f in findings if f.status == "PENDING")
    approved_count = sum(1 for f in findings if f.status == "APPROVED")
    rejected_count = sum(1 for f in findings if f.status == "REJECTED")
    deferred_count = sum(1 for f in findings if f.status == "DEFERRED")

    stats = [
        ("総指摘数", len(findings)),
        ("HIGH", high_count),
        ("MEDIUM", medium_count),
        ("LOW", low_count),
        ("", ""),
        ("未対応", pending_count),
        ("承認", approved_count),
        ("却下", rejected_count),
        ("保留", deferred_count),
    ]

    for label, value in stats:
        label_cell = ws.cell(row=row_num, column=1, value=label)
        label_cell.font = Font(bold=True)
        if label:
            label_cell.border = _THIN_BORDER
        value_cell = ws.cell(row=row_num, column=2, value=value)
        if label:
            value_cell.border = _THIN_BORDER
            # severity行に色をつける
            fill = _SEVERITY_FILLS.get(label)
            if fill:
                label_cell.fill = fill
                value_cell.fill = fill
        row_num += 1

    # エクスポート日時
    row_num += 1
    ws.cell(row=row_num, column=1, value="エクスポート日時").font = Font(
        italic=True, color="808080"
    )
    ws.cell(
        row=row_num,
        column=2,
        value=datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC"),
    ).font = Font(italic=True, color="808080")


def _build_findings_sheet(wb: Workbook, findings: list[ReviewFinding]) -> None:
    """Sheet2: 指摘事項一覧シートを構築する。"""
    ws = wb.create_sheet("指摘事項一覧")

    headers = [
        ("No.", 6),
        ("重要度", 10),
        ("種別", 15),
        ("箇所", 15),
        ("問題箇所テキスト", 30),
        ("問題内容", 40),
        ("改善提案", 40),
        ("根拠", 30),
        ("ステータス", 12),
        ("コメント", 30),
    ]

    # ヘッダー行
    for col, (name, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT
        cell.border = _THIN_BORDER
        ws.column_dimensions[cell.column_letter].width = width

    # データ行
    wrap_alignment = Alignment(vertical="top", wrap_text=True)
    for idx, finding in enumerate(findings, 1):
        row = idx + 1
        values = [
            idx,
            finding.severity,
            finding.issue_type,
            finding.location or "",
            finding.original_text or "",
            finding.description,
            finding.suggestion or "",
            finding.rationale or "",
            _STATUS_LABELS.get(finding.status, finding.status),
            finding.comment or "",
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = _THIN_BORDER
            cell.alignment = wrap_alignment

        # severity列に色をつける
        severity_fill = _SEVERITY_FILLS.get(finding.severity)
        if severity_fill:
            ws.cell(row=row, column=2).fill = severity_fill

    # オートフィルター設定
    if findings:
        ws.auto_filter.ref = f"A1:J{len(findings) + 1}"

    # ヘッダー行を固定
    ws.freeze_panes = "A2"


def _format_datetime(dt: datetime | None) -> str:
    """datetimeを表示用文字列に変換する。"""
    if dt is None:
        return ""
    return dt.strftime("%Y-%m-%d %H:%M")
