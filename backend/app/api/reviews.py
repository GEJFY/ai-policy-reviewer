"""
API endpoints for Review management.

このモジュールはAIレビュー機能のAPIエンドポイントを提供する。
レビューは非同期バックグラウンドタスクで実行され、進捗状況を追跡できる。

主要機能:
    - レビューの作成と実行開始
    - レビュー一覧・詳細の取得
    - 進捗状況のリアルタイム確認
    - レビュー結果の削除

レビューのライフサイクル:
    1. POST /reviews でレビュー作成（status: pending）
    2. バックグラウンドタスクで実行開始（status: processing）
    3. 各チェック項目を順次処理（check_item status: pending → processing → completed）
    4. 全チェック完了後、レビュー完了（status: completed）
    5. エラー発生時は失敗状態に移行（status: failed）

依存関係:
    - ReviewEngine: AIによるレビュー処理
    - Azure OpenAI: LLMによる分析
    - VectorStore: 類似用語検索
"""

from urllib.parse import quote

from fastapi import APIRouter, Depends, HTTPException, BackgroundTasks, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import case, func as sa_func
from sqlalchemy.orm import Session, joinedload
from typing import Optional

from app.db.database import get_db
from app.models.document import Document
from app.models.check_item import CheckItem
from app.models.review import Review, ReviewCheckItem, ReviewFinding
from app.schemas.review import (
    ReviewCreate,
    ReviewResponse,
    ReviewDetailResponse,
    ReviewCheckItemStatus,
    BatchReviewCreate,
    BatchReviewResponse,
)
from app.services.review_engine import review_engine
from app.services.docx_generator import generate_revised_docx
from app.core.logging_config import get_logger

# モジュール専用ロガー
logger = get_logger(__name__)

router = APIRouter(prefix="/api/v1/reviews", tags=["Reviews"])


@router.get("", response_model=list[ReviewDetailResponse])
async def list_reviews(
    skip: int = Query(default=0, ge=0),
    limit: int = Query(default=100, ge=1, le=1000),
    document_id: Optional[int] = Query(default=None),
    status: Optional[str] = Query(default=None),
    db: Session = Depends(get_db),
):
    """
    レビュー一覧を取得する。

    ページネーションとフィルタリングをサポートし、
    各レビューの指摘事項の集計情報を含む詳細レスポンスを返す。

    Args:
        skip: スキップするレコード数（デフォルト: 0）
        limit: 取得する最大レコード数（1-1000、デフォルト: 100）
        document_id: 特定の文書IDでフィルタリング（オプション）
        status: レビューステータスでフィルタリング
                有効値: pending, processing, completed, failed
        db: データベースセッション（依存性注入）

    Returns:
        list[ReviewDetailResponse]: レビュー詳細情報のリスト
            - id: レビューID
            - document_id: 対象文書ID
            - document_title: 文書タイトル
            - status: レビューステータス
            - finding_count: 総指摘件数
            - high_count: 高優先度指摘件数
            - medium_count: 中優先度指摘件数
            - low_count: 低優先度指摘件数
            - created_at: 作成日時
            - completed_at: 完了日時（未完了の場合null）

    Example:
        GET /api/v1/reviews?status=completed&limit=10
        → 完了済みレビューの最新10件を取得
    """
    query = db.query(Review)

    if document_id:
        query = query.filter(Review.document_id == document_id)

    if status:
        query = query.filter(Review.status == status)

    reviews = (
        query.options(joinedload(Review.document))
        .order_by(Review.created_at.desc())
        .offset(skip)
        .limit(limit)
        .all()
    )

    if not reviews:
        return []

    # 集計クエリ: 全レビューの severity 別 finding 件数を1クエリで取得
    review_ids = [r.id for r in reviews]
    counts = (
        db.query(
            ReviewFinding.review_id,
            sa_func.count(ReviewFinding.id).label("total"),
            sa_func.sum(case((ReviewFinding.severity == "HIGH", 1), else_=0)).label(
                "high"
            ),
            sa_func.sum(case((ReviewFinding.severity == "MEDIUM", 1), else_=0)).label(
                "medium"
            ),
            sa_func.sum(case((ReviewFinding.severity == "LOW", 1), else_=0)).label(
                "low"
            ),
        )
        .filter(ReviewFinding.review_id.in_(review_ids))
        .group_by(ReviewFinding.review_id)
        .all()
    )
    count_map = {
        row.review_id: {
            "total": row.total,
            "high": int(row.high or 0),
            "medium": int(row.medium or 0),
            "low": int(row.low or 0),
        }
        for row in counts
    }

    result = []
    for review in reviews:
        c = count_map.get(review.id, {"total": 0, "high": 0, "medium": 0, "low": 0})
        result.append(
            ReviewDetailResponse(
                id=review.id,
                document_id=review.document_id,
                status=review.status,
                created_at=review.created_at,
                completed_at=review.completed_at,
                document_title=review.document.title if review.document else None,
                finding_count=c["total"],
                high_count=c["high"],
                medium_count=c["medium"],
                low_count=c["low"],
            )
        )

    return result


@router.get("/{review_id}", response_model=ReviewDetailResponse)
async def get_review(review_id: int, db: Session = Depends(get_db)):
    """
    特定のレビューの詳細情報を取得する。

    指摘事項の集計とチェック項目ごとの進捗状況を含む
    詳細なレビュー情報を返す。フロントエンドのレビュー結果画面で使用。

    Args:
        review_id: 取得するレビューのID
        db: データベースセッション（依存性注入）

    Returns:
        ReviewDetailResponse: レビュー詳細情報
            - check_items: 各チェック項目の進捗状況リスト
            - finding_count: 総指摘件数
            - high/medium/low_count: 重要度別指摘件数

    Raises:
        HTTPException(404): 指定IDのレビューが存在しない場合

    Note:
        レビュー実行中（status=processing）の場合、
        check_items.statusで各チェック項目の進捗を確認できる。
    """
    review = (
        db.query(Review)
        .options(joinedload(Review.document))
        .filter(Review.id == review_id)
        .first()
    )
    if not review:
        raise HTTPException(status_code=404, detail="Review not found")

    # チェック項目ステータスを1クエリで取得（JOINで名前も取得）
    check_items_rows = (
        db.query(ReviewCheckItem.check_item_id, CheckItem.name, ReviewCheckItem.status)
        .join(CheckItem, CheckItem.id == ReviewCheckItem.check_item_id)
        .filter(ReviewCheckItem.review_id == review_id)
        .all()
    )
    check_items_status = [
        ReviewCheckItemStatus(
            check_item_id=row.check_item_id,
            check_item_name=row.name,
            status=row.status,
        )
        for row in check_items_rows
    ]

    # Finding件数を集計クエリで取得（全件ロード不要）
    counts = (
        db.query(
            sa_func.count(ReviewFinding.id).label("total"),
            sa_func.sum(case((ReviewFinding.severity == "HIGH", 1), else_=0)).label(
                "high"
            ),
            sa_func.sum(case((ReviewFinding.severity == "MEDIUM", 1), else_=0)).label(
                "medium"
            ),
            sa_func.sum(case((ReviewFinding.severity == "LOW", 1), else_=0)).label(
                "low"
            ),
        )
        .filter(ReviewFinding.review_id == review_id)
        .first()
    )

    return ReviewDetailResponse(
        id=review.id,
        document_id=review.document_id,
        status=review.status,
        created_at=review.created_at,
        completed_at=review.completed_at,
        document_title=review.document.title if review.document else None,
        check_items=check_items_status,
        finding_count=counts.total if counts else 0,
        high_count=int(counts.high or 0) if counts else 0,
        medium_count=int(counts.medium or 0) if counts else 0,
        low_count=int(counts.low or 0) if counts else 0,
    )


@router.post("", response_model=ReviewResponse, status_code=201)
async def create_review(
    request: ReviewCreate,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
):
    """
    新規レビューを作成し、バックグラウンドで実行を開始する。

    このエンドポイントはレビュー作成後、即座にレスポンスを返す。
    実際のAIレビュー処理はバックグラウンドタスクとして非同期実行される。

    処理フロー:
        1. 文書の存在確認とOCR完了チェック
        2. 指定されたチェック項目の存在確認
        3. レビューレコードの作成（status: pending）
        4. ReviewCheckItem関連付けの作成
        5. ReviewEngineの利用可能性確認
        6. バックグラウンドタスクへの登録
        7. レスポンス返却（この時点でHTTP 201）

    Args:
        request: レビュー作成リクエスト
            - document_id: レビュー対象の文書ID
            - check_item_ids: 実行するチェック項目IDのリスト
        background_tasks: FastAPIバックグラウンドタスクハンドラ
        db: データベースセッション（依存性注入）

    Returns:
        ReviewResponse: 作成されたレビュー情報（status=pending）

    Raises:
        HTTPException(404): 指定された文書が存在しない
        HTTPException(400): 文書のOCRが未完了、またはチェック項目が存在しない
        HTTPException(503): ReviewEngine（Azure OpenAI）が利用不可

    バックグラウンド処理:
        - execute_review_task関数が非同期で実行される
        - 処理中はstatus=processingに更新される
        - 完了時にstatus=completed、失敗時にstatus=failedに更新
        - 進捗確認は GET /{review_id}/status で可能

    Example:
        POST /api/v1/reviews
        {
            "document_id": 1,
            "check_item_ids": [1, 2, 3]
        }
        → レビュー作成後、AIが用語・文法・構成をチェック

    Note:
        大規模文書の場合、レビュー完了まで数分かかる場合がある。
        フロントエンドはstatusエンドポイントでポーリング監視を推奨。
    """
    # Validate document exists and is ready
    document = db.query(Document).filter(Document.id == request.document_id).first()
    if not document:
        raise HTTPException(status_code=404, detail="Document not found")

    if document.ocr_status != "completed":
        raise HTTPException(
            status_code=400,
            detail=f"Document OCR not completed. Current status: {document.ocr_status}",
        )

    # Validate check items exist
    for check_item_id in request.check_item_ids:
        check_item = db.query(CheckItem).filter(CheckItem.id == check_item_id).first()
        if not check_item:
            raise HTTPException(
                status_code=400, detail=f"Check item {check_item_id} not found"
            )

    # Create review
    review = Review(
        document_id=request.document_id,
        status="pending",
    )
    db.add(review)
    db.commit()
    db.refresh(review)

    # Create review check item associations
    for check_item_id in request.check_item_ids:
        review_check = ReviewCheckItem(
            review_id=review.id,
            check_item_id=check_item_id,
            status="pending",
        )
        db.add(review_check)
    db.commit()

    # Check if review engine is available
    if not review_engine.is_available():
        review.status = "failed"
        db.commit()
        raise HTTPException(
            status_code=503,
            detail="Review engine not available. Check LLM provider configuration.",
        )

    # Schedule review execution
    background_tasks.add_task(
        execute_review_task,
        review.id,
        request.check_item_ids,
    )

    return review


@router.get("/{review_id}/status")
async def get_review_status(review_id: int, db: Session = Depends(get_db)):
    """
    レビューの進捗状況をリアルタイムで取得する。

    フロントエンドからのポーリングに使用。
    各チェック項目の処理状況と全体の進捗率を返す。

    Args:
        review_id: 確認するレビューのID
        db: データベースセッション（依存性注入）

    Returns:
        dict: 進捗情報
            - status: レビュー全体のステータス
            - total_checks: 総チェック項目数
            - completed_checks: 完了したチェック項目数
            - processing_checks: 処理中のチェック項目数
            - progress_percent: 進捗率（0-100%）

    Raises:
        HTTPException(404): 指定IDのレビューが存在しない場合

    Example Response:
        {
            "status": "processing",
            "total_checks": 7,
            "completed_checks": 3,
            "processing_checks": 1,
            "progress_percent": 42.86
        }

    Note:
        フロントエンドでは1-2秒間隔でのポーリングを推奨。
        status=completed または status=failed でポーリング終了。
    """
    review = db.query(Review).filter(Review.id == review_id).first()
    if not review:
        raise HTTPException(status_code=404, detail="Review not found")

    # Get check items progress
    check_items = (
        db.query(ReviewCheckItem).filter(ReviewCheckItem.review_id == review_id).all()
    )
    total = len(check_items)
    completed = sum(1 for ci in check_items if ci.status == "completed")
    processing = sum(1 for ci in check_items if ci.status == "processing")

    return {
        "status": review.status,
        "total_checks": total,
        "completed_checks": completed,
        "processing_checks": processing,
        "progress_percent": (completed / total * 100) if total > 0 else 0,
    }


@router.delete("/{review_id}", status_code=204)
async def delete_review(review_id: int, db: Session = Depends(get_db)):
    """
    レビューと関連する全てのデータを削除する。

    ReviewFindingおよびReviewCheckItemはカスケード削除される。
    実行中のレビューも削除可能だが、バックグラウンド処理は
    次のDB操作時にレビュー不在を検知して自動終了する。

    Args:
        review_id: 削除するレビューのID
        db: データベースセッション（依存性注入）

    Returns:
        None: 204 No Contentを返す

    Raises:
        HTTPException(404): 指定IDのレビューが存在しない場合

    Caution:
        この操作は取り消し不可。指摘事項と承認履歴も全て削除される。
    """
    review = db.query(Review).filter(Review.id == review_id).first()
    if not review:
        raise HTTPException(status_code=404, detail="Review not found")

    db.delete(review)
    db.commit()
    return None


@router.post("/batch", response_model=BatchReviewResponse, status_code=201)
async def create_batch_review(
    request: BatchReviewCreate,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
):
    """Create reviews for multiple documents at once."""
    # Validate check items
    for check_item_id in request.check_item_ids:
        check_item = db.query(CheckItem).filter(CheckItem.id == check_item_id).first()
        if not check_item:
            raise HTTPException(
                status_code=400, detail=f"Check item {check_item_id} not found"
            )

    if not review_engine.is_available():
        raise HTTPException(
            status_code=503,
            detail="Review engine not available. Check LLM provider configuration.",
        )

    created_reviews = []
    failed_document_ids = []

    for doc_id in request.document_ids:
        document = db.query(Document).filter(Document.id == doc_id).first()
        if not document:
            failed_document_ids.append(doc_id)
            continue
        if document.ocr_status != "completed":
            failed_document_ids.append(doc_id)
            continue

        review = Review(document_id=doc_id, status="pending")
        db.add(review)
        db.commit()
        db.refresh(review)

        for check_item_id in request.check_item_ids:
            review_check = ReviewCheckItem(
                review_id=review.id,
                check_item_id=check_item_id,
                status="pending",
            )
            db.add(review_check)
        db.commit()

        background_tasks.add_task(
            execute_review_task, review.id, request.check_item_ids
        )
        created_reviews.append(review)

    return BatchReviewResponse(
        created_reviews=created_reviews,
        failed_document_ids=failed_document_ids,
    )


@router.get("/{review_id}/revised-document")
async def download_revised_document(review_id: int, db: Session = Depends(get_db)):
    """Download revised document as DOCX with approved suggestions applied."""
    review = db.query(Review).filter(Review.id == review_id).first()
    if not review:
        raise HTTPException(status_code=404, detail="Review not found")

    document = db.query(Document).filter(Document.id == review.document_id).first()
    if not document or not document.extracted_text:
        raise HTTPException(status_code=404, detail="Document text not available")

    # Get approved findings
    approved_findings = (
        db.query(ReviewFinding)
        .filter(
            ReviewFinding.review_id == review_id,
            ReviewFinding.status == "APPROVED",
            ReviewFinding.original_text.isnot(None),
        )
        .all()
    )

    # Apply replacements
    revised_text = document.extracted_text
    changes_applied = 0
    replacements = []

    for f in approved_findings:
        suggestion = getattr(f, "edited_suggestion", None) or f.suggestion
        if not suggestion or not f.original_text:
            continue
        idx = revised_text.find(f.original_text)
        if idx >= 0:
            replacements.append(
                {
                    "start": idx,
                    "end": idx + len(f.original_text),
                    "replacement": suggestion,
                }
            )

    replacements.sort(key=lambda r: r["start"], reverse=True)
    for r in replacements:
        revised_text = (
            revised_text[: r["start"]] + r["replacement"] + revised_text[r["end"] :]
        )
        changes_applied += 1

    # Generate DOCX
    docx_buffer = generate_revised_docx(
        title=document.title,
        original_text=document.extracted_text,
        revised_text=revised_text,
        changes_applied=changes_applied,
        total_approved=len(approved_findings),
    )

    # Build filename
    base_name = (
        document.title.rsplit(".", 1)[0] if "." in document.title else document.title
    )
    filename = f"{base_name}_改訂版.docx"
    encoded_filename = quote(filename)

    return StreamingResponse(
        docx_buffer,
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
        },
    )


@router.post("/bulk-export")
async def bulk_export_reviews(
    request: dict,
    db: Session = Depends(get_db),
):
    """Export multiple reviews as a single Excel file."""
    from app.api.export import _build_workbook, _encode_filename
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    review_ids = request.get("review_ids", [])
    if not review_ids:
        raise HTTPException(status_code=400, detail="review_ids is required")

    # Fetch all reviews with their documents and findings
    reviews = (
        db.query(Review)
        .options(joinedload(Review.document))
        .filter(Review.id.in_(review_ids))
        .order_by(Review.created_at.desc())
        .all()
    )

    if not reviews:
        raise HTTPException(status_code=404, detail="No reviews found")

    # If single review, use existing single-export logic
    if len(reviews) == 1:
        review = reviews[0]
        document = review.document
        findings = (
            db.query(ReviewFinding)
            .filter(ReviewFinding.review_id == review.id)
            .order_by(ReviewFinding.severity.desc(), ReviewFinding.created_at)
            .all()
        )
        wb = _build_workbook(review, document, findings)
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        doc_title = document.title if document else f"review_{review.id}"
        safe_title = doc_title.replace("/", "_").replace("\\", "_").replace(":", "_")
        filename = f"{safe_title}_レビュー結果.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{_encode_filename(filename)}"
            },
        )

    # Multiple reviews: build a workbook with one sheet per review
    wb = Workbook()
    wb.remove(wb.active)  # type: ignore[arg-type]

    for review in reviews:
        findings = (
            db.query(ReviewFinding)
            .filter(ReviewFinding.review_id == review.id)
            .order_by(ReviewFinding.severity.desc(), ReviewFinding.created_at)
            .all()
        )
        doc_title = review.document.title if review.document else f"Review_{review.id}"
        # Sheet names max 31 chars
        sheet_name = doc_title[:28] + "..." if len(doc_title) > 31 else doc_title
        # Remove invalid sheet name characters
        for ch in ["\\", "/", "*", "?", ":", "[", "]"]:
            sheet_name = sheet_name.replace(ch, "_")

        ws = wb.create_sheet(title=sheet_name)
        # Build findings table directly
        from app.api.export import (
            _HEADER_FONT,
            _HEADER_FILL,
            _HEADER_ALIGNMENT,
            _THIN_BORDER,
            _SEVERITY_FILLS,
            _STATUS_LABELS,
        )

        headers = [
            ("No.", 6),
            ("重要度", 10),
            ("種別", 15),
            ("箇所", 15),
            ("問題箇所テキスト", 30),
            ("問題内容", 40),
            ("改善提案", 40),
            ("根拠", 30),
            ("ステータス", 12),
            ("コメント", 30),
        ]

        for col, (name, width) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=name)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _HEADER_ALIGNMENT
            cell.border = _THIN_BORDER
            ws.column_dimensions[cell.column_letter].width = width

        wrap_alignment = Alignment(vertical="top", wrap_text=True)
        for idx, finding in enumerate(findings, 1):
            row = idx + 1
            values = [
                idx,
                finding.severity,
                finding.issue_type,
                finding.location or "",
                finding.original_text or "",
                finding.description,
                finding.suggestion or "",
                finding.rationale or "",
                _STATUS_LABELS.get(finding.status, finding.status),
                finding.comment or "",
            ]
            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.border = _THIN_BORDER
                cell.alignment = wrap_alignment
            severity_fill = _SEVERITY_FILLS.get(finding.severity)
            if severity_fill:
                ws.cell(row=row, column=2).fill = severity_fill

        if findings:
            ws.auto_filter.ref = f"A1:J{len(findings) + 1}"
        ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = "レビュー結果_一括.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{_encode_filename(filename)}"
        },
    )


async def execute_review_task(review_id: int, check_item_ids: list[int]):
    """
    バックグラウンドでレビューを実行するタスク関数。

    FastAPIのBackgroundTasksから呼び出され、メインリクエストとは
    独立してレビュー処理を実行する。新しいDBセッションを作成し、
    処理完了後に必ずクローズする。

    処理フロー:
        1. 新しいDBセッションを作成
        2. ReviewEngineのexecute_reviewを呼び出し
        3. 各チェック項目を順次処理（LLM API呼び出し）
        4. 指摘事項をDBに保存
        5. 完了時にstatus=completedに更新
        6. DBセッションをクローズ

    エラーハンドリング:
        - 例外発生時はレビューをfailed状態に更新
        - エラー詳細はログに記録される
        - DBセッションは必ずfinally節でクローズ

    Args:
        review_id: 実行するレビューのID
        check_item_ids: 実行するチェック項目IDのリスト

    Note:
        この関数はHTTPレスポンス後に非同期で実行される。
        進捗確認はGET /reviews/{id}/statusを使用。

    Warning:
        長時間実行（数分）になる可能性がある。
        Azure OpenAI APIのRate Limit対応は
        ReviewEngine内で実装されている。
    """
    import asyncio
    from app.db.database import SessionLocal

    # レビュータスクのタイムアウト（10分）
    REVIEW_TIMEOUT_SECONDS = 600

    logger.info(
        f"Starting background review task: review_id={review_id}, check_items={check_item_ids}"
    )

    db = SessionLocal()
    try:
        await asyncio.wait_for(
            review_engine.execute_review(
                db=db,
                review_id=review_id,
                check_item_ids=check_item_ids,
            ),
            timeout=REVIEW_TIMEOUT_SECONDS,
        )
        logger.info(f"Review task completed successfully: review_id={review_id}")
    except asyncio.TimeoutError:
        logger.error(
            f"Review task timed out after {REVIEW_TIMEOUT_SECONDS}s: review_id={review_id}"
        )
        review = db.query(Review).filter(Review.id == review_id).first()
        if review:
            review.status = "failed"
            db.commit()
    except Exception as e:
        logger.error(
            f"Review execution failed: review_id={review_id}, error={e}", exc_info=True
        )
        review = db.query(Review).filter(Review.id == review_id).first()
        if review:
            review.status = "failed"
            db.commit()
    finally:
        db.close()
