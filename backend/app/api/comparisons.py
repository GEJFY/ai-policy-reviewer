"""API endpoints for parent-subsidiary policy comparison."""

import io
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from urllib.parse import quote

from app.db.database import get_db
from app.models.document import Document
from app.models.comparison import (
    ComparisonProject,
    ComparisonCheckItem,
    ComparisonResult,
)
from app.schemas.comparison import (
    ComparisonProjectCreate,
    SetSubsidiaryRequest,
    ChecklistEditRequest,
    ComparisonProjectResponse,
    ComparisonProjectDetailResponse,
    ComparisonCheckItemResponse,
    ComparisonResultResponse,
)
from app.services.llm_service import llm_service

router = APIRouter(prefix="/api/v1/comparisons", tags=["Comparisons"])


def _build_project_response(project: ComparisonProject) -> dict:
    """Build response dict from a ComparisonProject."""
    return {
        "id": project.id,
        "name": project.name,
        "description": project.description,
        "parent_document_id": project.parent_document_id,
        "parent_document_title": (
            project.parent_document.title if project.parent_document else ""
        ),
        "subsidiary_document_id": project.subsidiary_document_id,
        "subsidiary_document_title": (
            project.subsidiary_document.title if project.subsidiary_document else None
        ),
        "status": project.status,
        "check_item_count": len(project.check_items),
        "result_count": len(project.results),
        "created_at": project.created_at,
        "updated_at": project.updated_at,
    }


@router.get("", response_model=list[ComparisonProjectResponse])
async def list_projects(db: Session = Depends(get_db)):
    """List all comparison projects."""
    projects = (
        db.query(ComparisonProject).order_by(ComparisonProject.created_at.desc()).all()
    )
    return [_build_project_response(p) for p in projects]


@router.post("", response_model=ComparisonProjectDetailResponse, status_code=201)
async def create_project(
    request: ComparisonProjectCreate, db: Session = Depends(get_db)
):
    """Create a new comparison project."""
    parent = (
        db.query(Document).filter(Document.id == request.parent_document_id).first()
    )
    if not parent:
        raise HTTPException(status_code=404, detail="Parent document not found")

    project = ComparisonProject(
        name=request.name,
        description=request.description,
        parent_document_id=request.parent_document_id,
        status="created",
    )
    db.add(project)
    db.commit()
    db.refresh(project)

    resp = _build_project_response(project)
    resp["check_items"] = []
    resp["results"] = []
    return resp


@router.get("/{project_id}", response_model=ComparisonProjectDetailResponse)
async def get_project(project_id: int, db: Session = Depends(get_db)):
    """Get a comparison project with details."""
    project = (
        db.query(ComparisonProject).filter(ComparisonProject.id == project_id).first()
    )
    if not project:
        raise HTTPException(status_code=404, detail="Comparison project not found")

    resp = _build_project_response(project)
    resp["check_items"] = [
        ComparisonCheckItemResponse(
            id=ci.id,
            item_text=ci.item_text,
            category=ci.category,
            order_index=ci.order_index,
        )
        for ci in project.check_items
    ]
    resp["results"] = [
        ComparisonResultResponse(
            id=r.id,
            check_item_id=r.check_item_id,
            check_item_text=r.check_item.item_text if r.check_item else "",
            status=r.status,
            parent_text=r.parent_text,
            subsidiary_text=r.subsidiary_text,
            explanation=r.explanation,
        )
        for r in project.results
    ]
    return resp


@router.delete("/{project_id}", status_code=204)
async def delete_project(project_id: int, db: Session = Depends(get_db)):
    """Delete a comparison project."""
    project = (
        db.query(ComparisonProject).filter(ComparisonProject.id == project_id).first()
    )
    if not project:
        raise HTTPException(status_code=404, detail="Comparison project not found")
    db.delete(project)
    db.commit()
    return None


@router.post("/{project_id}/generate-checklist")
async def generate_checklist(project_id: int, db: Session = Depends(get_db)):
    """Generate checklist items from the parent document using LLM."""
    from app.services.comparison_service import generate_checklist as gen_checklist

    project = (
        db.query(ComparisonProject).filter(ComparisonProject.id == project_id).first()
    )
    if not project:
        raise HTTPException(status_code=404, detail="Comparison project not found")

    if not llm_service.is_available():
        raise HTTPException(
            status_code=503,
            detail="LLM service not available. Check provider configuration.",
        )

    items = await gen_checklist(db, project_id)

    # Clear existing check items
    db.query(ComparisonCheckItem).filter(
        ComparisonCheckItem.project_id == project_id
    ).delete()

    for i, item in enumerate(items):
        ci = ComparisonCheckItem(
            project_id=project_id,
            item_text=item.get("item_text", ""),
            category=item.get("category"),
            order_index=i,
        )
        db.add(ci)

    project.status = "checklist_ready"  # type: ignore[assignment]
    db.commit()

    return {"message": f"{len(items)} checklist items generated", "count": len(items)}


@router.put("/{project_id}/checklist")
async def update_checklist(
    project_id: int,
    request: ChecklistEditRequest,
    db: Session = Depends(get_db),
):
    """Update checklist items (user edits after generation)."""
    project = (
        db.query(ComparisonProject).filter(ComparisonProject.id == project_id).first()
    )
    if not project:
        raise HTTPException(status_code=404, detail="Comparison project not found")

    db.query(ComparisonCheckItem).filter(
        ComparisonCheckItem.project_id == project_id
    ).delete()

    for i, item in enumerate(request.items):
        ci = ComparisonCheckItem(
            project_id=project_id,
            item_text=item.item_text,
            category=item.category,
            order_index=i,
        )
        db.add(ci)

    db.commit()
    return {"message": f"{len(request.items)} checklist items updated"}


@router.put("/{project_id}/subsidiary")
async def set_subsidiary(
    project_id: int,
    request: SetSubsidiaryRequest,
    db: Session = Depends(get_db),
):
    """Set the subsidiary document for comparison."""
    project = (
        db.query(ComparisonProject).filter(ComparisonProject.id == project_id).first()
    )
    if not project:
        raise HTTPException(status_code=404, detail="Comparison project not found")

    sub_doc = (
        db.query(Document).filter(Document.id == request.subsidiary_document_id).first()
    )
    if not sub_doc:
        raise HTTPException(status_code=404, detail="Subsidiary document not found")

    project.subsidiary_document_id = request.subsidiary_document_id  # type: ignore[assignment]
    db.commit()

    return {"message": "Subsidiary document set"}


@router.post("/{project_id}/compare")
async def run_comparison(project_id: int, db: Session = Depends(get_db)):
    """Run comparison between parent and subsidiary documents."""
    from app.services.comparison_service import compare_single_item

    project = (
        db.query(ComparisonProject).filter(ComparisonProject.id == project_id).first()
    )
    if not project:
        raise HTTPException(status_code=404, detail="Comparison project not found")

    if not project.subsidiary_document_id:
        raise HTTPException(status_code=400, detail="Subsidiary document not set")

    if not project.check_items:
        raise HTTPException(
            status_code=400, detail="No checklist items. Generate checklist first."
        )

    if not llm_service.is_available():
        raise HTTPException(
            status_code=503,
            detail="LLM service not available. Check provider configuration.",
        )

    parent_doc = (
        db.query(Document).filter(Document.id == project.parent_document_id).first()
    )
    subsidiary_doc = (
        db.query(Document).filter(Document.id == project.subsidiary_document_id).first()
    )
    if not parent_doc or not subsidiary_doc:
        raise HTTPException(status_code=404, detail="Document not found")

    project.status = "comparing"  # type: ignore[assignment]
    db.commit()

    # Clear existing results
    db.query(ComparisonResult).filter(
        ComparisonResult.project_id == project_id
    ).delete()
    db.commit()

    results: list[dict[str, str]] = []
    for ci in project.check_items:
        try:
            result = await compare_single_item(
                db, str(ci.item_text), parent_doc, subsidiary_doc
            )
            cr = ComparisonResult(
                project_id=project_id,
                check_item_id=ci.id,
                status=result["status"],
                parent_text=result.get("parent_text"),
                subsidiary_text=result.get("subsidiary_text"),
                explanation=result.get("explanation"),
            )
            db.add(cr)
            results.append(result)
        except Exception as e:
            cr = ComparisonResult(
                project_id=project_id,
                check_item_id=ci.id,
                status="DIFFERENT",
                explanation=f"比較処理エラー: {str(e)}",
            )
            db.add(cr)
            results.append({"status": "DIFFERENT", "explanation": str(e)})

    project.status = "completed"  # type: ignore[assignment]
    db.commit()

    status_counts: dict[str, int] = {}
    for r in results:
        s = r["status"]
        status_counts[s] = status_counts.get(s, 0) + 1

    return {
        "message": f"Comparison completed for {len(results)} items",
        "total": len(results),
        "status_counts": status_counts,
    }


@router.get("/{project_id}/export")
async def export_results(project_id: int, db: Session = Depends(get_db)):
    """Export comparison results as Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    project = (
        db.query(ComparisonProject).filter(ComparisonProject.id == project_id).first()
    )
    if not project:
        raise HTTPException(status_code=404, detail="Comparison project not found")

    wb = openpyxl.Workbook()

    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "概要"
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    summary_data: list[list[object]] = [
        ["プロジェクト名", project.name],
        ["説明", project.description or ""],
        [
            "親会社規程",
            project.parent_document.title if project.parent_document else "",
        ],
        [
            "子会社規程",
            (project.subsidiary_document.title if project.subsidiary_document else ""),
        ],
        ["ステータス", project.status],
        ["チェック項目数", len(project.check_items)],
        ["比較結果数", len(project.results)],
    ]
    for row_idx, row_data in enumerate(summary_data, 1):
        label, value = row_data[0], row_data[1]
        cell_label = ws_summary.cell(row=row_idx, column=1, value=label)
        cell_label.font = Font(bold=True)
        cell_label.border = thin_border
        cell_value = ws_summary.cell(row=row_idx, column=2, value=value)
        cell_value.border = thin_border

    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 50

    # Status summary
    status_counts: dict[str, int] = {}
    for r in project.results:
        s = str(r.status)
        status_counts[s] = status_counts.get(s, 0) + 1

    row_start = len(summary_data) + 2
    ws_summary.cell(row=row_start, column=1, value="判定別集計").font = Font(
        bold=True, size=12
    )

    STATUS_LABELS = {
        "COMPLIANT": "適合",
        "STRICTER": "より厳格",
        "LOOSER": "緩い",
        "MISSING": "欠落",
        "DIFFERENT": "異なる",
    }
    STATUS_COLORS = {
        "COMPLIANT": "C6EFCE",
        "STRICTER": "BDD7EE",
        "LOOSER": "FCE4D6",
        "MISSING": "FFC7CE",
        "DIFFERENT": "FFF2CC",
    }

    for i, (status, label) in enumerate(STATUS_LABELS.items()):
        row = row_start + 1 + i
        cell_s = ws_summary.cell(row=row, column=1, value=label)
        cell_s.border = thin_border
        cell_s.fill = PatternFill(
            start_color=STATUS_COLORS.get(status, "FFFFFF"),
            end_color=STATUS_COLORS.get(status, "FFFFFF"),
            fill_type="solid",
        )
        cell_c = ws_summary.cell(row=row, column=2, value=status_counts.get(status, 0))
        cell_c.border = thin_border

    # Results sheet
    ws_results = wb.create_sheet("比較結果")
    headers = [
        "#",
        "カテゴリ",
        "チェック項目",
        "判定",
        "親会社規程",
        "子会社規程",
        "説明",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws_results.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row_idx, result in enumerate(project.results, 2):
        ci = result.check_item
        data = [
            row_idx - 1,
            ci.category if ci else "",
            ci.item_text if ci else "",
            STATUS_LABELS.get(result.status, result.status),
            result.parent_text or "",
            result.subsidiary_text or "",
            result.explanation or "",
        ]
        for col, val in enumerate(data, 1):
            cell = ws_results.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if col == 4:
                cell.fill = PatternFill(
                    start_color=STATUS_COLORS.get(result.status, "FFFFFF"),
                    end_color=STATUS_COLORS.get(result.status, "FFFFFF"),
                    fill_type="solid",
                )

    ws_results.column_dimensions["A"].width = 5
    ws_results.column_dimensions["B"].width = 15
    ws_results.column_dimensions["C"].width = 40
    ws_results.column_dimensions["D"].width = 12
    ws_results.column_dimensions["E"].width = 40
    ws_results.column_dimensions["F"].width = 40
    ws_results.column_dimensions["G"].width = 40

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"comparison_{project.name}.xlsx"
    encoded_filename = quote(filename)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
        },
    )
