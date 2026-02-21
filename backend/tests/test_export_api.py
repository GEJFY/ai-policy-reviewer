"""
Tests for Review Export API endpoint.

レビュー結果のExcelエクスポートAPIのテスト。
"""

import io

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.main import app
from app.db.database import get_db
from app.models.base import Base
from app.models.document import Document
from app.models.review import Review, ReviewFinding

SQLALCHEMY_DATABASE_URL = "sqlite:///:memory:"
engine = create_engine(
    SQLALCHEMY_DATABASE_URL,
    connect_args={"check_same_thread": False},
    poolclass=StaticPool,
)
TestingSessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)


@pytest.fixture(autouse=True)
def setup_database():
    Base.metadata.create_all(bind=engine)
    yield
    Base.metadata.drop_all(bind=engine)


@pytest.fixture
def db_session():
    session = TestingSessionLocal()
    try:
        yield session
    finally:
        session.close()


@pytest.fixture
def client(db_session):
    def override_get_db():
        try:
            yield db_session
        finally:
            pass

    app.dependency_overrides[get_db] = override_get_db
    with TestClient(app) as test_client:
        yield test_client
    app.dependency_overrides.clear()


@pytest.fixture
def sample_review_with_findings(db_session):
    """レビューと指摘事項のテストデータを作成"""
    doc = Document(
        title="情報セキュリティポリシー.pdf",
        file_path="/tmp/test.pdf",
        ocr_status="completed",
    )
    db_session.add(doc)
    db_session.commit()
    db_session.refresh(doc)

    review = Review(document_id=doc.id, status="completed")
    db_session.add(review)
    db_session.commit()
    db_session.refresh(review)

    findings = [
        ReviewFinding(
            review_id=review.id,
            issue_type="TERMINOLOGY",
            severity="HIGH",
            description="「社員」は正式用語「従業員」に統一すべき",
            location="第3条",
            original_text="社員は所定の手続きに従い申請する",
            suggestion="「従業員」に変更",
            rationale="用語辞書の定義に基づく",
            status="APPROVED",
            comment="修正済み",
        ),
        ReviewFinding(
            review_id=review.id,
            issue_type="GRAMMAR",
            severity="MEDIUM",
            description="「等」の使用が曖昧",
            location="第5条第2項",
            original_text="書類等を提出する",
            suggestion="具体的な書類名を列挙",
            status="PENDING",
        ),
        ReviewFinding(
            review_id=review.id,
            issue_type="COMPLIANCE",
            severity="LOW",
            description="参照先の条文番号が不明確",
            location="第10条",
            original_text="関連法令に従う",
            suggestion="具体的な法令名と条文番号を記載",
            status="DEFERRED",
            comment="次回改訂時に対応",
        ),
    ]
    for f in findings:
        db_session.add(f)
    db_session.commit()
    for f in findings:
        db_session.refresh(f)

    return {"review": review, "findings": findings, "document": doc}


class TestExportReviewExcel:
    """GET /api/v1/reviews/{review_id}/export のテスト"""

    def test_export_success(self, client: TestClient, sample_review_with_findings):
        """Excelエクスポートが正常に完了"""
        review_id = sample_review_with_findings["review"].id
        response = client.get(f"/api/v1/reviews/{review_id}/export")
        assert response.status_code == 200
        assert (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            in response.headers["content-type"]
        )
        assert "attachment" in response.headers["content-disposition"]

    def test_export_filename_contains_document_title(
        self, client: TestClient, sample_review_with_findings
    ):
        """ファイル名に文書タイトルが含まれる"""
        review_id = sample_review_with_findings["review"].id
        response = client.get(f"/api/v1/reviews/{review_id}/export")
        disposition = response.headers["content-disposition"]
        # UTF-8エンコードされた日本語ファイル名が含まれる
        assert "UTF-8''" in disposition
        assert ".xlsx" in disposition

    def test_export_review_not_found(self, client: TestClient):
        """存在しないレビューIDで404"""
        response = client.get("/api/v1/reviews/9999/export")
        assert response.status_code == 404

    def test_export_has_two_sheets(
        self, client: TestClient, sample_review_with_findings
    ):
        """Excelファイルに2シートが存在"""
        review_id = sample_review_with_findings["review"].id
        response = client.get(f"/api/v1/reviews/{review_id}/export")
        wb = load_workbook(io.BytesIO(response.content))
        assert len(wb.sheetnames) == 2
        assert wb.sheetnames[0] == "レビュー概要"
        assert wb.sheetnames[1] == "指摘事項一覧"

    def test_export_summary_sheet_content(
        self, client: TestClient, sample_review_with_findings
    ):
        """概要シートにレビュー情報が含まれる"""
        review_id = sample_review_with_findings["review"].id
        response = client.get(f"/api/v1/reviews/{review_id}/export")
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb["レビュー概要"]

        # タイトル
        assert ws["A1"].value == "レビュー結果レポート"

        # 文書名
        assert ws["A3"].value == "文書名"
        assert ws["B3"].value == "情報セキュリティポリシー.pdf"

        # ステータス
        assert ws["A5"].value == "ステータス"
        assert ws["B5"].value == "completed"

    def test_export_summary_sheet_statistics(
        self, client: TestClient, sample_review_with_findings
    ):
        """概要シートに正確な統計値が含まれる"""
        review_id = sample_review_with_findings["review"].id
        response = client.get(f"/api/v1/reviews/{review_id}/export")
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb["レビュー概要"]

        # 総指摘数のセルを探す
        total_found = False
        for row in ws.iter_rows(min_row=1, max_row=30, max_col=2, values_only=False):
            if row[0].value == "総指摘数":
                assert row[1].value == 3
                total_found = True
            elif row[0].value == "HIGH":
                assert row[1].value == 1
            elif row[0].value == "MEDIUM":
                assert row[1].value == 1
            elif row[0].value == "LOW":
                assert row[1].value == 1
        assert total_found, "総指摘数のセルが見つからない"

    def test_export_findings_sheet_content(
        self, client: TestClient, sample_review_with_findings
    ):
        """指摘事項シートにデータ行が含まれる"""
        review_id = sample_review_with_findings["review"].id
        response = client.get(f"/api/v1/reviews/{review_id}/export")
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb["指摘事項一覧"]

        # ヘッダー行
        assert ws.cell(row=1, column=1).value == "No."
        assert ws.cell(row=1, column=2).value == "重要度"
        assert ws.cell(row=1, column=6).value == "問題内容"

        # データ行（3件）
        assert ws.cell(row=2, column=1).value == 1  # No.
        assert ws.cell(row=4, column=1).value == 3  # 最後の行

        # 5行目（データなし）
        assert ws.cell(row=5, column=1).value is None

    def test_export_findings_sheet_severity_values(
        self, client: TestClient, sample_review_with_findings
    ):
        """指摘事項シートのseverity値が正確"""
        review_id = sample_review_with_findings["review"].id
        response = client.get(f"/api/v1/reviews/{review_id}/export")
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb["指摘事項一覧"]

        severities = []
        for row in range(2, 5):
            severities.append(ws.cell(row=row, column=2).value)

        assert "HIGH" in severities
        assert "MEDIUM" in severities
        assert "LOW" in severities

    def test_export_findings_sheet_status_labels(
        self, client: TestClient, sample_review_with_findings
    ):
        """ステータスが日本語ラベルで表示される"""
        review_id = sample_review_with_findings["review"].id
        response = client.get(f"/api/v1/reviews/{review_id}/export")
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb["指摘事項一覧"]

        statuses = []
        for row in range(2, 5):
            statuses.append(ws.cell(row=row, column=9).value)

        assert "承認" in statuses
        assert "未対応" in statuses
        assert "保留" in statuses

    def test_export_empty_findings(self, client: TestClient, db_session):
        """指摘事項が0件のレビューでもエクスポート可能"""
        doc = Document(
            title="空レビュー.pdf",
            file_path="/tmp/empty.pdf",
            ocr_status="completed",
        )
        db_session.add(doc)
        db_session.commit()
        review = Review(document_id=doc.id, status="completed")
        db_session.add(review)
        db_session.commit()

        response = client.get(f"/api/v1/reviews/{review.id}/export")
        assert response.status_code == 200

        wb = load_workbook(io.BytesIO(response.content))
        ws = wb["指摘事項一覧"]
        # ヘッダーのみ、データ行なし
        assert ws.cell(row=1, column=1).value == "No."
        assert ws.cell(row=2, column=1).value is None

    def test_export_valid_excel_file(
        self, client: TestClient, sample_review_with_findings
    ):
        """レスポンスが有効なExcelファイル"""
        review_id = sample_review_with_findings["review"].id
        response = client.get(f"/api/v1/reviews/{review_id}/export")

        # openpyxlで正常に開けることを確認
        wb = load_workbook(io.BytesIO(response.content))
        assert wb is not None
        wb.close()

    def test_export_auto_filter_applied(
        self, client: TestClient, sample_review_with_findings
    ):
        """指摘事項シートにオートフィルターが設定されている"""
        review_id = sample_review_with_findings["review"].id
        response = client.get(f"/api/v1/reviews/{review_id}/export")
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb["指摘事項一覧"]
        assert ws.auto_filter.ref is not None

    def test_export_freeze_panes(self, client: TestClient, sample_review_with_findings):
        """指摘事項シートのヘッダー行が固定されている"""
        review_id = sample_review_with_findings["review"].id
        response = client.get(f"/api/v1/reviews/{review_id}/export")
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb["指摘事項一覧"]
        assert ws.freeze_panes == "A2"
