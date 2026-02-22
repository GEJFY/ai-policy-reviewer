"""
Tests for bulk export API endpoint.

レビュー一括エクスポートAPIのテスト。
"""

import io

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.main import app
from app.db.database import get_db
from app.models.base import Base
from app.models.document import Document
from app.models.review import Review, ReviewFinding

SQLALCHEMY_DATABASE_URL = "sqlite:///:memory:"
engine = create_engine(
    SQLALCHEMY_DATABASE_URL,
    connect_args={"check_same_thread": False},
    poolclass=StaticPool,
)
TestingSessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)


@pytest.fixture(autouse=True)
def setup_database():
    Base.metadata.create_all(bind=engine)
    yield
    Base.metadata.drop_all(bind=engine)


@pytest.fixture
def db_session():
    session = TestingSessionLocal()
    try:
        yield session
    finally:
        session.close()


@pytest.fixture
def client(db_session):
    def override_get_db():
        try:
            yield db_session
        finally:
            pass

    app.dependency_overrides[get_db] = override_get_db
    with TestClient(app) as test_client:
        yield test_client
    app.dependency_overrides.clear()


def _create_review_with_findings(db_session, title: str, finding_count: int = 2):
    """テスト用レビューデータを作成するヘルパー"""
    doc = Document(
        title=f"{title}.pdf",
        file_path=f"/tmp/{title}.pdf",
        ocr_status="completed",
    )
    db_session.add(doc)
    db_session.commit()
    db_session.refresh(doc)

    review = Review(document_id=doc.id, status="completed")
    db_session.add(review)
    db_session.commit()
    db_session.refresh(review)

    for i in range(finding_count):
        finding = ReviewFinding(
            review_id=review.id,
            issue_type="TERMINOLOGY",
            severity=["HIGH", "MEDIUM", "LOW"][i % 3],
            description=f"テスト指摘{i + 1}",
            location=f"第{i + 1}条",
            original_text=f"問題テキスト{i + 1}",
            suggestion=f"修正提案{i + 1}",
            status="PENDING",
        )
        db_session.add(finding)
    db_session.commit()

    return review


class TestBulkExport:
    """POST /api/v1/reviews/bulk-export のテスト"""

    def test_bulk_export_single_review(self, client: TestClient, db_session):
        """単一レビューのエクスポートが正常動作"""
        review = _create_review_with_findings(db_session, "テスト文書")

        response = client.post(
            "/api/v1/reviews/bulk-export",
            json={"review_ids": [review.id]},
        )
        assert response.status_code == 200
        assert (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            in response.headers["content-type"]
        )

        # 有効なExcelファイル
        wb = load_workbook(io.BytesIO(response.content))
        assert wb is not None
        wb.close()

    def test_bulk_export_multiple_reviews(self, client: TestClient, db_session):
        """複数レビューのエクスポートが正常動作"""
        review1 = _create_review_with_findings(db_session, "セキュリティポリシー", 3)
        review2 = _create_review_with_findings(db_session, "就業規則", 2)

        response = client.post(
            "/api/v1/reviews/bulk-export",
            json={"review_ids": [review1.id, review2.id]},
        )
        assert response.status_code == 200

        wb = load_workbook(io.BytesIO(response.content))
        # 各レビューに1シートずつ
        assert len(wb.sheetnames) == 2
        wb.close()

    def test_bulk_export_sheet_names(self, client: TestClient, db_session):
        """シート名が文書タイトルに基づく"""
        review1 = _create_review_with_findings(db_session, "文書A", 1)
        review2 = _create_review_with_findings(db_session, "文書B", 1)

        response = client.post(
            "/api/v1/reviews/bulk-export",
            json={"review_ids": [review1.id, review2.id]},
        )
        wb = load_workbook(io.BytesIO(response.content))
        assert "文書A.pdf" in wb.sheetnames
        assert "文書B.pdf" in wb.sheetnames
        wb.close()

    def test_bulk_export_sheet_content(self, client: TestClient, db_session):
        """各シートにヘッダーとデータが含まれる"""
        review = _create_review_with_findings(db_session, "内容確認用", 2)

        # 2件以上で bulk sheet モードにする
        review2 = _create_review_with_findings(db_session, "ダミー", 1)
        response = client.post(
            "/api/v1/reviews/bulk-export",
            json={"review_ids": [review.id, review2.id]},
        )
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb["内容確認用.pdf"]

        # ヘッダー行
        assert ws.cell(row=1, column=1).value == "No."
        assert ws.cell(row=1, column=2).value == "重要度"
        assert ws.cell(row=1, column=6).value == "問題内容"

        # データ行
        assert ws.cell(row=2, column=1).value == 1
        assert ws.cell(row=3, column=1).value == 2
        wb.close()

    def test_bulk_export_freeze_panes(self, client: TestClient, db_session):
        """ヘッダー行が固定されている"""
        review1 = _create_review_with_findings(db_session, "FP確認1", 1)
        review2 = _create_review_with_findings(db_session, "FP確認2", 1)

        response = client.post(
            "/api/v1/reviews/bulk-export",
            json={"review_ids": [review1.id, review2.id]},
        )
        wb = load_workbook(io.BytesIO(response.content))
        for ws in wb.worksheets:
            assert ws.freeze_panes == "A2"
        wb.close()

    def test_bulk_export_empty_ids(self, client: TestClient):
        """空の review_ids で400"""
        response = client.post(
            "/api/v1/reviews/bulk-export",
            json={"review_ids": []},
        )
        assert response.status_code == 400

    def test_bulk_export_nonexistent_ids(self, client: TestClient):
        """存在しないIDで404"""
        response = client.post(
            "/api/v1/reviews/bulk-export",
            json={"review_ids": [99999]},
        )
        assert response.status_code == 404

    def test_bulk_export_auto_filter(self, client: TestClient, db_session):
        """指摘事項があるシートにオートフィルターが設定されている"""
        review1 = _create_review_with_findings(db_session, "AF確認1", 2)
        review2 = _create_review_with_findings(db_session, "AF確認2", 1)

        response = client.post(
            "/api/v1/reviews/bulk-export",
            json={"review_ids": [review1.id, review2.id]},
        )
        wb = load_workbook(io.BytesIO(response.content))
        for ws in wb.worksheets:
            assert ws.auto_filter.ref is not None
        wb.close()
